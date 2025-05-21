import sys
from tkinter import Widget
from PyQt5.QtWidgets import QApplication, QMainWindow ,QTableWidgetItem
from Ui_girisekrani import Ui_MainWindow as UiGirisEkrani
from Ui_soruekleme import Ui_MainWindow as UiSoruEkleme
from Ui_secmeekrani import Ui_MainWindow as UiSoruSecme
from PyQt5.QtWidgets import QListWidgetItem, QFileDialog
from PyQt5.QtWidgets import QFileDialog ,QTableWidgetItem
from openpyxl import Workbook , load_workbook
from PyQt5.QtPrintSupport import QPrinter
from PyQt5.QtGui import QTextDocument

    
# Soru Ekleme Penceresi
class SoruEklemePenceresi(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = UiSoruEkleme()
        self.ui.setupUi(self)





class SoruEklemePenceresi(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = UiSoruEkleme()
        self.ui.setupUi(self)
        
        self.ui.pushButton_ekle.clicked.connect(self.soru_ekle)
        self.ui.pushButton_2.clicked.connect(self.excel_kaydet)

    def soru_ekle(self):
        soru = self.ui.textEdit_soru.toPlainText()
        secenekler = [
            self.ui.lineEdit_a.text(),
            self.ui.lineEdit_b.text(),
            self.ui.lineEdit_c.text(),
            self.ui.lineEdit_d.text(),
            self.ui.lineEdit_e.text()
        ]
        dogru = ''
        if self.ui.radio_a.isChecked(): dogru = 'A'
        elif self.ui.radio_b.isChecked(): dogru = 'B'
        elif self.ui.radio_c.isChecked(): dogru = 'C'
        elif self.ui.radio_d.isChecked(): dogru = 'D'
        elif self.ui.radio_e.isChecked(): dogru = 'E'
        
        metin = f"{soru} | {' | '.join(secenekler)} | Doğru: {dogru}"
        self.ui.listWidget_sorular.addItem(QListWidgetItem(metin))

    def excel_kaydet(self):
        dosya_adi, _ = QFileDialog.getSaveFileName(self, "Excel olarak kaydet", "", "Excel Files (*.xlsx)")
        if dosya_adi:
            wb = Workbook()
            ws = wb.active
            ws.append(["SORU", "A", "B", "C", "D", "E", "DOĞRU"])
            for i in range(self.ui.listWidget_sorular.count()):
                item = self.ui.listWidget_sorular.item(i).text()
                parcali = item.split(" | ")
                soru = parcali[0]
                secenekler = parcali[1:6]
                dogru = parcali[6].replace("Doğru: ", "")
                ws.append([soru] + secenekler + [dogru])
            wb.save(dosya_adi)

# Soru Seçme Penceresi
class SoruSecmePenceresi(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = UiSoruSecme()
        self.ui.setupUi(self)


# Giriş Ekranı (Ana Menü)
class GirisPenceresi(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = UiGirisEkrani()
        self.ui.setupUi(self)

        # Alt pencereler
        self.soru_ekleme_penceresi = None
        self.soru_secme_penceresi = None

        # Butonlara tıklanınca yapılacak işlemler
        self.ui.pushButton.clicked.connect(self.soru_ekle_ac)
        self.ui.pushButton_2.clicked.connect(self.soru_secme_ac)

    def soru_ekle_ac(self):
        if self.soru_ekleme_penceresi is None or not self.soru_ekleme_penceresi.isVisible():
            self.soru_ekleme_penceresi = SoruEklemePenceresi()
            self.soru_ekleme_penceresi.show()
        else:
            self.soru_ekleme_penceresi.raise_()
            self.soru_ekleme_penceresi.activateWindow()

    def soru_secme_ac(self):
        if self.soru_secme_penceresi is None or not self.soru_secme_penceresi.isVisible():
            self.soru_secme_penceresi = SoruSecmePenceresi()
            self.soru_secme_penceresi.show()
        else:
            self.soru_secme_penceresi.raise_()
            self.soru_secme_penceresi.activateWindow()





class SoruSecmePenceresi(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = UiSoruSecme()
        self.ui.setupUi(self)
        
        self.ui.pushButton.clicked.connect(self.dosyayi_yukle)
        self.ui.pushButton_2.clicked.connect(self.yazdir)

    def dosyayi_yukle(self):
        dosya_adi, _ = QFileDialog.getOpenFileName(self, "Excel Dosyası Seç", "", "Excel Files (*.xlsx)")
        if dosya_adi:
            wb = load_workbook(dosya_adi)
            ws = wb.active
            self.ui.tableWidget.setRowCount(0)
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                self.ui.tableWidget.insertRow(i)
                for j, value in enumerate(row):
                    self.ui.tableWidget.setItem(i, j, QTableWidgetItem(str(value)))

    def yazdir(self):
        dosya_adi, _ = QFileDialog.getSaveFileName(self, "PDF olarak yazdır", "", "PDF Files (*.pdf)")
        if dosya_adi:
            if not dosya_adi.endswith(".pdf"):
                dosya_adi += ".pdf"

            html = "<h2>Sorular Tablosu</h2><table border='1' cellspacing='0' cellpadding='4'>"
            html += "<tr><th>SORU</th><th>1.SEÇENEK</th><th>2.SEÇENEK</th><th>3.SEÇENEK</th><th>4.SEÇENEK</th><th>5.SEÇENEK</th><th>DOĞRU SEÇENEK</th></tr>"

            for row in range(self.ui.tableWidget.rowCount()):
                html += "<tr>"
                for col in range(self.ui.tableWidget.columnCount()):
                    item = self.ui.tableWidget.item(row, col)
                    html += f"<td>{item.text() if item else ''}</td>"
                html += "</tr>"

            html += "</table>"

            doc = QTextDocument()
            doc.setHtml(html)

            printer = QPrinter()
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(dosya_adi)

            doc.print_(printer)
# Uygulama başlat
if __name__ == "__main__":
    app = QApplication(sys.argv)
    giris = GirisPenceresi()
    giris.show()
    sys.exit(app.exec_())